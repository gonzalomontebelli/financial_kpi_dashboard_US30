import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import LineChart, Reference
from openpyxl.utils import get_column_letter

from config import OUTPUT_DIR, EXCEL_DIR

def _style_kpi_cell(cell, is_title=False):
    if is_title:
        cell.font = Font(bold=True, size=12)
    else:
        cell.font = Font(bold=True, size=18)
    cell.alignment = Alignment(horizontal="center", vertical="center")

def main():
    EXCEL_DIR.mkdir(parents=True, exist_ok=True)

    kpis = pd.read_csv(OUTPUT_DIR / "kpis_summary.csv")
    rolling = pd.read_csv(OUTPUT_DIR / "rolling_volatility.csv")
    monthly = pd.read_csv(OUTPUT_DIR / "monthly_returns.csv")
    yearly = pd.read_csv(OUTPUT_DIR / "yearly_returns.csv")

    # Convert dates
    rolling["date"] = pd.to_datetime(rolling["date"])
    monthly["month"] = pd.to_datetime(monthly["month"])
    yearly["year"] = pd.to_datetime(yearly["year"])

    wb = Workbook()
    ws_data = wb.active
    ws_data.title = "data"

    ws_calc = wb.create_sheet("calc")
    ws_dash = wb.create_sheet("dashboard")

    # --- data sheet (date, close) ---
    ws_data.append(["date", "close"])
    for _, r in rolling[["date", "close"]].iterrows():
        ws_data.append([r["date"].date(), float(r["close"])])

    # Column widths
    ws_data.column_dimensions["A"].width = 14
    ws_data.column_dimensions["B"].width = 14
    ws_data.freeze_panes = "A2"

    # --- calc sheet (rolling vol + returns) ---
    cols = ["date","close","daily_return","log_return","drawdown","vol_20d_ann","vol_60d_ann"]
    ws_calc.append(cols)
    for _, r in rolling[cols].iterrows():
        ws_calc.append([
            r["date"].date(),
            float(r["close"]),
            None if pd.isna(r["daily_return"]) else float(r["daily_return"]),
            None if pd.isna(r["log_return"]) else float(r["log_return"]),
            None if pd.isna(r["drawdown"]) else float(r["drawdown"]),
            None if pd.isna(r["vol_20d_ann"]) else float(r["vol_20d_ann"]),
            None if pd.isna(r["vol_60d_ann"]) else float(r["vol_60d_ann"]),
        ])
    ws_calc.freeze_panes = "A2"
    for i, w in enumerate([14,14,16,16,14,14,14], start=1):
        ws_calc.column_dimensions[get_column_letter(i)].width = w

    # --- dashboard sheet ---
    ws_dash["A1"] = "DJIA KPI Dashboard"
    ws_dash["A1"].font = Font(bold=True, size=20)
    ws_dash.merge_cells("A1:H1")

    # KPI cards
    # Layout: titles row 3, values row 4 (each merged across 2 columns)
    k = kpis.iloc[0].to_dict()
    cards = [
        ("As of", k["as_of_date"]),
        ("Last Close", k["last_close"]),
        ("Daily Return", k["last_daily_return"]),
        ("YTD Return", k["ytd_return"]),
        ("1Y Return", k["one_year_return"]),
        ("Vol 20D (ann.)", k["vol_20d_ann"]),
        ("Vol 60D (ann.)", k["vol_60d_ann"]),
        ("Max Drawdown", k["max_drawdown"]),
        ("Sharpe (simple, ann.)", k["sharpe_simple_ann"]),
    ]

    start_col = 1
    row_title = 3
    row_value = 4
    col = start_col
    fill = PatternFill("solid", fgColor="F2F2F2")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # 3 rows of 3 cards (9 total)
    positions = [
        ("A","B"), ("C","D"), ("E","F"),
        ("A","B"), ("C","D"), ("E","F"),
        ("A","B"), ("C","D"), ("E","F"),
    ]
    rows = [3,3,3, 6,6,6, 9,9,9]

    for idx, ((c1,c2), r0) in enumerate(zip(positions, rows)):
        title, val = cards[idx]
        # title
        ws_dash.merge_cells(f"{c1}{r0}:{c2}{r0}")
        cell_t = ws_dash[f"{c1}{r0}"]
        cell_t.value = title
        _style_kpi_cell(cell_t, is_title=True)
        cell_t.fill = fill
        cell_t.border = border

        # value
        ws_dash.merge_cells(f"{c1}{r0+1}:{c2}{r0+1}")
        cell_v = ws_dash[f"{c1}{r0+1}"]
        cell_v.value = val
        _style_kpi_cell(cell_v, is_title=False)
        cell_v.fill = PatternFill("solid", fgColor="FFFFFF")
        cell_v.border = border

    # Format numbers
    # Identify cells by title
    num_formats = {
        "Last Close": "#,##0.00",
        "Daily Return": "0.00%",
        "YTD Return": "0.00%",
        "1Y Return": "0.00%",
        "Vol 20D (ann.)": "0.00%",
        "Vol 60D (ann.)": "0.00%",
        "Max Drawdown": "0.00%",
        "Sharpe (simple, ann.)": "0.00",
    }
    for idx, ((c1,c2), r0) in enumerate(zip(positions, rows)):
        title, _ = cards[idx]
        val_cell = ws_dash[f"{c1}{r0+1}"]
        if title in num_formats:
            val_cell.number_format = num_formats[title]

    # Charts
    # Price chart from calc sheet
    chart1 = LineChart()
    chart1.title = "Close Price"
    chart1.y_axis.title = "Price"
    chart1.x_axis.title = "Date"

    data_ref = Reference(ws_calc, min_col=2, min_row=1, max_row=ws_calc.max_row)  # close
    dates_ref = Reference(ws_calc, min_col=1, min_row=2, max_row=ws_calc.max_row)
    chart1.add_data(data_ref, titles_from_data=True)
    chart1.set_categories(dates_ref)
    chart1.height = 8
    chart1.width = 24
    ws_dash.add_chart(chart1, "A12")

    # Vol chart
    chart2 = LineChart()
    chart2.title = "Rolling Volatility (Annualized)"
    chart2.y_axis.title = "Vol"
    chart2.x_axis.title = "Date"
    vol_ref = Reference(ws_calc, min_col=6, min_row=1, max_col=7, max_row=ws_calc.max_row)
    chart2.add_data(vol_ref, titles_from_data=True)
    chart2.set_categories(dates_ref)
    chart2.height = 8
    chart2.width = 24
    ws_dash.add_chart(chart2, "A30")

    # Monthly return chart (bar-like via line chart; simpler)
    ws_calc2 = wb.create_sheet("monthly")
    ws_calc2.append(["month","month_return"])
    for _, r in monthly.iterrows():
        ws_calc2.append([r["month"].date(), float(r["month_return"])])
    ws_calc2.freeze_panes = "A2"
    ws_calc2.column_dimensions["A"].width = 14
    ws_calc2.column_dimensions["B"].width = 16

    chart3 = LineChart()
    chart3.title = "Monthly Returns"
    m_data = Reference(ws_calc2, min_col=2, min_row=1, max_row=ws_calc2.max_row)
    m_dates = Reference(ws_calc2, min_col=1, min_row=2, max_row=ws_calc2.max_row)
    chart3.add_data(m_data, titles_from_data=True)
    chart3.set_categories(m_dates)
    chart3.y_axis.title = "Return"
    chart3.height = 8
    chart3.width = 24
    ws_dash.add_chart(chart3, "A48")

    # Freeze panes
    ws_dash.freeze_panes = "A2"

    out_path = EXCEL_DIR / "DJI_KPI_Dashboard.xlsx"
    wb.save(out_path)
    print(f"OK. Excel dashboard written to: {out_path}")

if __name__ == "__main__":
    main()
